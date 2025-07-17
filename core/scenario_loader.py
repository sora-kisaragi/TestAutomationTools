import openpyxl
from typing import Any, Dict, List, Optional

class ScenarioLoader:
    def __init__(self, scenario_path: str):
        self.scenario_path = scenario_path

    def extract_scenarios_from_sheet(self, ws) -> Optional[Dict[str, Any]]:
        """
        指定シートから画面名・シナリオ・テスト項目を抽出
        戻り値: {
            "screen_name": str,
            "scenarios": [
                {
                    "name": str,  # シナリオ名
                    "testitems": [ {...}, ... ]  # テスト項目リスト
                }, ...
            ]
        }
        """
        # 画面名の取得
        a1 = ws["A1"].value if ws["A1"].value else ""
        b1 = ws["B1"].value if ws["B1"].value else ""
        a2 = ws["A2"].value if ws["A2"].value else ""
        if a1 in ("テスト画面名", "画面名", "画面"):
            screen_name = a2 if a2 else b1
        else:
            screen_name = a1
        if not screen_name:
            screen_name = ws.title
        print(f"[LOG] 画面名: {screen_name}")

        scenarios = []
        row_idx = 1
        max_row = ws.max_row
        while row_idx <= max_row:
            a_val = ws.cell(row=row_idx, column=1).value
            # A列に"シナリオ名"が現れたら、その直下がシナリオ名
            if a_val == "シナリオ名":
                scenario_name = ws.cell(row=row_idx+1, column=1).value
                print(f"[LOG] シナリオ名検出: {scenario_name} (row={row_idx+1})")
                scenario = {"name": scenario_name, "testitems": []}
                # さらに2行下がNoで始まる場合はテスト項目ヘッダー
                header_row = row_idx + 2
                if ws.cell(row=header_row, column=1).value == "No":
                    # ヘッダー取得
                    headers = []
                    for col in range(1, ws.max_column + 1):
                        header = ws.cell(row=header_row, column=col).value
                        if header is None:
                            break
                        headers.append(header)
                    print(f"[LOG] テスト項目ヘッダー: {headers} (row={header_row})")
                    # テスト項目データ抽出
                    data_row = header_row + 1
                    while data_row <= max_row:
                        first_col = ws.cell(row=data_row, column=1).value
                        # 次のシナリオ名や空行でbreak
                        if first_col == "シナリオ名" or (first_col is None and ws.cell(row=data_row, column=2).value is None):
                            print(f"[LOG] テスト項目抽出終了 row={data_row}")
                            break
                        # データ行を辞書化
                        testitem = {}
                        for i, header in enumerate(headers):
                            testitem[header] = ws.cell(row=data_row, column=i+1).value
                        print(f"[LOG] テスト項目抽出: {testitem}")
                        scenario["testitems"].append(testitem)
                        data_row += 1
                    scenarios.append(scenario)
                    row_idx = data_row
                    continue
                else:
                    print(f"[WARN] テスト項目ヘッダーが見つかりません row={header_row}")
                    row_idx += 1
                    continue
            row_idx += 1
        return {
            "screen_name": screen_name,
            "scenarios": scenarios
        }

    def load_all_scenarios(self) -> List[Dict[str, Any]]:
        """
        全シートからシナリオデータを抽出し整形
        戻り値: [ { ... }, ... ]
        """
        wb = openpyxl.load_workbook(self.scenario_path, data_only=True)
        exclude_prefixes = ("不具合報告", "共通項目", "サンプル")
        all_data = []
        for sheetname in wb.sheetnames:
            if any(sheetname.startswith(prefix) for prefix in exclude_prefixes):
                continue
            ws = wb[sheetname]
            data = self.extract_scenarios_from_sheet(ws)
            if data and data["scenarios"]:
                all_data.append(data)
        return all_data

# --- mainブロック ---
if __name__ == "__main__":
    path = input("Excelファイルのパス : ")
    loader = ScenarioLoader(path)
    all_scenarios = loader.load_all_scenarios()
    # ページ(画面)ごとのシナリオ・テストケース・テスト項目数を表示
    print("\nページ(画面)ごとのシナリオ数:")
    for page in all_scenarios:
        print(f"- {page['screen_name']} ({len(page['scenarios'])}件)")
        for s in page["scenarios"]:
            print(f"  - {s['name']} (テスト項目: {len(s['testitems'])}件)")
